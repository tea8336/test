# coding:utf-8
# excel.py
# yang.wenbo


import os
import openpyxl

from datetime import date
from dateutil import parser


class Excel:
    'Excel类'

    def __init__(self):
        '【初始化】'
        self.str_path = os.path.dirname(os.path.abspath(__file__))
        self.office_excel = openpyxl.load_workbook('%s\\files\\btc.xlsx' % self.str_path)
        self.dict_years = {}

    def update_excel(self):
        '【处理Excel数据】'
        # 打开源数据
        excel_sheet = self.office_excel['btc']
        # 数据分组
        self.data_group(excel_sheet)
        # 写入sheet
        self.create_sheet()
        # 保存
        self.excel_save('%s\\files\\study.xlsx' % self.str_path)
        print('保存成功')

    def data_group(self, excel_sheet: object):
        '【数据分组】'
        # TODO 遍历数据，按年份分组
        for data_row in excel_sheet.rows:
            # 标题列不处理
            if data_row[0].value.strip() == '日期':
                continue
            # 日期格式转换
            date_year = parser.parse(data_row[0].value).strftime('%Y')
            # 数据分组
            if self.dict_years.get(date_year) == None:
                self.dict_years[date_year] = [[data_row[0].value, data_row[1].value]]
            else:
                self.dict_years[date_year].append([data_row[0].value, data_row[1].value])

    def create_sheet(self):
        '【写入sheet】'
        for str_year in self.dict_years.keys():
            # 标题列
            excel_new_sheet = self.office_excel.create_sheet(str_year)
            excel_new_sheet['A1'] = '日期'
            excel_new_sheet['B1'] = '价格'
            # 数据
            int_rows = 2
            int_len_a = 0
            for list_data in self.dict_years[str_year]:
                excel_new_sheet[f'A{int_rows}'] = list_data[0]
                excel_new_sheet[f'B{int_rows}'] = list_data[1]
                if len(list_data[0]) > int_len_a:
                    int_len_a = len(list_data[0])
                int_rows += 1
            # 平均值
            excel_new_sheet[f'A{int_rows}'] = '平均值'
            # excel_new_sheet[f'B{int_rows}'] = f'=average(B2:B{int_rows- 1})'
            excel_new_sheet[f'B{int_rows}'] = ('=average(B2:B%i)' % (int_rows-1))

            # 列宽
            excel_new_sheet.column_dimensions['A'].width = int_len_a + 2
            excel_new_sheet.column_dimensions['B'].width = 16

    def excel_save(self, str_name: str):
        '【保存】'
        self.office_excel.save(str_name)


def main():
    excel_obj = Excel()
    excel_obj.update_excel()

if __name__ == '__main__':
    main()
